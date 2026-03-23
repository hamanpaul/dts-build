from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


def read_table_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(path)
    raise ValueError(f"unsupported table format: {path}")


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        rows: list[dict[str, str]] = []
        for row in reader:
            rows.append({str(k): "" if v is None else str(v).strip() for k, v in row.items()})
        return rows


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index = _detect_header_index(rows)
    headers = _normalize_headers(rows[header_index])
    parsed_rows: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        parsed_rows.append(
            {
                headers[index]: "" if value is None else str(value).strip()
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
        )
    return parsed_rows


def _detect_header_index(rows: list[tuple[object, ...]]) -> int:
    for index, row in enumerate(rows[:15]):
        non_empty = ["" if value is None else str(value).strip() for value in row]
        filled = [value for value in non_empty if value]
        lowered = " ".join(filled).lower()
        if len(filled) >= 4 and ("pin" in lowered or "note" in lowered or "active" in lowered):
            return index
    return 0


def _normalize_headers(row: tuple[object, ...]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        header = "" if value is None else str(value).strip()
        if not header:
            header = f"column_{index}"
        count = seen.get(header, 0) + 1
        seen[header] = count
        if count > 1:
            header = f"{header}_{count}"
        headers.append(header)
    return headers
