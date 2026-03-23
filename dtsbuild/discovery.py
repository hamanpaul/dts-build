from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .manifest import FAMILY_DEFAULTS, write_manifest


@dataclass
class DiscoveryResult:
    folder: Path
    project: str
    family: str
    spreadsheet_files: list[Path] = field(default_factory=list)
    pdf_files: list[Path] = field(default_factory=list)
    dts_files: list[Path] = field(default_factory=list)
    other_files: list[Path] = field(default_factory=list)
    suggested_manifest: dict[str, Any] = field(default_factory=dict)


def discover_folder(folder: Path) -> DiscoveryResult:
    project = folder.name.removeprefix("dtsin_") or folder.name

    files = sorted([path for path in folder.iterdir() if path.is_file()])
    spreadsheets = [path for path in files if path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}]
    pdfs = [path for path in files if path.suffix.lower() == ".pdf"]
    dts_files = [path for path in files if path.suffix.lower() == ".dts"]
    others = [path for path in files if path not in spreadsheets and path not in pdfs and path not in dts_files and path.name != "manifest.yaml"]

    family = _guess_family(spreadsheets)
    defaults = FAMILY_DEFAULTS.get(family, {})
    artifacts: dict[str, Any] = {}
    notes: list[str] = []

    classified_tables, unknown_tables = _classify_spreadsheets(folder, spreadsheets)
    artifacts.update(classified_tables)
    classified_dts = _classify_dts_files(folder, dts_files)
    artifacts.update(classified_dts)

    if pdfs:
        artifacts["schematic_pdfs"] = [path.relative_to(folder).as_posix() for path in pdfs]
    if unknown_tables:
        notes.append(f"Unclassified tables detected: {', '.join(unknown_tables)}")
    if not spreadsheets:
        notes.append("No spreadsheet files were found in the folder.")

    suggested_manifest = {
        "project": project,
        "family": family or "bcm68575",
        "profile": "unknownprofile",
        "refboard": "unknownrefboard",
        "model": project,
        "output_dts": f"{project}.dts",
        "output_dir": f"dtsout_{project}",
        "base_include": defaults.get("base_include", "inc/68375.dtsi"),
        "compatible": defaults.get("compatible", "brcm,bcm968375"),
        "artifacts": artifacts,
        "notes": [
            "Bootstrap-generated manifest. Please verify profile/refboard and artifact mappings."
        ]
        + notes,
    }

    return DiscoveryResult(
        folder=folder,
        project=project,
        family=family or "bcm68575",
        spreadsheet_files=spreadsheets,
        pdf_files=pdfs,
        dts_files=dts_files,
        other_files=others,
        suggested_manifest=suggested_manifest,
    )


def format_discovery(result: DiscoveryResult) -> str:
    lines = [
        f"Project: {result.project}",
        f"Family guess: {result.family or '<unknown>'}",
        "",
        "Detected spreadsheets:",
    ]
    if result.spreadsheet_files:
        lines.extend([f"  - {path.name}" for path in result.spreadsheet_files])
    else:
        lines.append("  - <none>")

    lines.extend(["", "Detected PDFs:"])
    if result.pdf_files:
        lines.extend([f"  - {path.name}" for path in result.pdf_files])
    else:
        lines.append("  - <none>")

    lines.extend(["", "Detected DTS files:"])
    if result.dts_files:
        lines.extend([f"  - {path.name}" for path in result.dts_files])
    else:
        lines.append("  - <none>")

    if result.other_files:
        lines.extend(["", "Other files:"])
        lines.extend([f"  - {path.name}" for path in result.other_files])

    lines.extend(
        [
            "",
            "No manifest.yaml was found.",
            "Suggested next step: run `python -m dtsbuild bootstrap-manifest <folder>` and then edit profile/refboard.",
        ]
    )
    return "\n".join(lines)


def bootstrap_manifest(folder: Path, force: bool = False) -> Path:
    manifest_path = folder / "manifest.yaml"
    if manifest_path.exists() and not force:
        raise FileExistsError(f"manifest already exists: {manifest_path}")

    discovery = discover_folder(folder)
    write_manifest(manifest_path, discovery.suggested_manifest)
    return manifest_path


def _guess_family(spreadsheets: list[Path]) -> str:
    for spreadsheet in spreadsheets:
        guessed = _guess_family_from_sheet(spreadsheet)
        if guessed:
            return guessed
    return ""


def _guess_family_from_sheet(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        name = path.name.lower()
        if "68575" in name or "68375" in name:
            return "bcm68575"
        return ""

    workbook = load_workbook(path, read_only=True, data_only=True)
    for ws in workbook.worksheets:
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            joined = " ".join("" if value is None else str(value) for value in row).lower()
            if "68575" in joined or "68375" in joined or "bcm68575" in joined:
                return "bcm68575"
            if row_index >= 12:
                break
    return ""


def _classify_spreadsheets(folder: Path, spreadsheets: list[Path]) -> tuple[dict[str, Any], list[str]]:
    artifacts: dict[str, Any] = {}
    unknown_tables: list[str] = []

    for path in spreadsheets:
        rel_path = path.relative_to(folder).as_posix()
        lowered = path.name.lower()
        if "gpio" in lowered or "led" in lowered:
            artifacts.setdefault("gpio_led_table", rel_path)
        elif "block" in lowered or "interface" in lowered:
            artifacts.setdefault("blockdiag_table", rel_path)
        elif "network" in lowered or "port" in lowered:
            artifacts.setdefault("network_table", rel_path)
        elif "ddr" in lowered or "memory" in lowered:
            artifacts.setdefault("ddr_table", rel_path)
        else:
            unknown_tables.append(path.name)
    return artifacts, unknown_tables


def _classify_dts_files(folder: Path, dts_files: list[Path]) -> dict[str, Any]:
    artifacts: dict[str, Any] = {}
    for path in dts_files:
        lowered = path.name.lower()
        if "ref" in lowered and "public_ref_dts" not in artifacts:
            artifacts["public_ref_dts"] = path.relative_to(folder).as_posix()
    return artifacts
