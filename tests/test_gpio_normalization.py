from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from dtsbuild.generator import generate_dts
from dtsbuild.manifest import write_manifest


class GpioNormalizationTest(unittest.TestCase):
    def test_generate_dts_normalizes_realistic_gpio_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir) / "dtsin_BGW720"
            folder.mkdir()
            (folder / "out").mkdir()

            workbook = Workbook()
            ws = workbook.active
            ws["D2"] = "BCM68575"
            ws["A4"] = "BCM68575公版 Pinout"
            ws["D4"] = "Pin"
            ws["E4"] = "Function Description"
            ws["G4"] = "I/O"
            ws["H4"] = "Active"
            ws["I4"] = "Note"
            ws["A5"] = "WAN_SFP_RX_LOS"
            ws["D5"] = "GPIO_03"
            ws["E5"] = "WAN_SFP_RX_LOS"
            ws["G5"] = "I"
            ws["H5"] = "Low"
            ws["I5"] = "LOW = Signal Detect"
            workbook.save(folder / "GPIO.xlsx")

            write_manifest(
                folder / "manifest.yaml",
                {
                    "project": "BGW720",
                    "family": "bcm68575",
                    "profile": "unknownprofile",
                    "refboard": "unknownrefboard",
                    "model": "BGW720",
                    "output_dts": "BGW720.dts",
                    "base_include": "inc/68375.dtsi",
                    "compatible": "brcm,bcm968375",
                    "artifacts": {"gpio_led_table": "GPIO.xlsx"},
                },
            )

            output = generate_dts(folder, backend="manual")
            content = output.read_text(encoding="utf-8")
            self.assertIn("gpio:WAN_SFP_RX_LOS", content)
            self.assertIn("pin_or_gpio=GPIO_03", content)


if __name__ == "__main__":
    unittest.main()
