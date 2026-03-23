from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from dtsbuild.discovery import bootstrap_manifest, discover_folder
from dtsbuild.manifest import load_manifest


class DiscoveryTest(unittest.TestCase):
    def test_discover_folder_and_bootstrap_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir) / "dtsin_BGW720"
            folder.mkdir()

            workbook = Workbook()
            ws = workbook.active
            ws["D2"] = "BCM68575"
            ws["A4"] = "BCM68575公版 Pinout"
            workbook.save(folder / "GPIO_R0A_20250730.xlsx")

            (folder / "MAINBOARD.pdf").write_bytes(b"%PDF-1.4\n")
            (folder / "DAUGHTER.pdf").write_bytes(b"%PDF-1.4\n")
            (folder / "968575REF1.dts").write_text('/dts-v1/;\n/ { model = "968575REF1"; };\n', encoding="utf-8")

            result = discover_folder(folder)
            self.assertEqual(result.project, "BGW720")
            self.assertEqual(result.family, "bcm68575")
            self.assertEqual(result.suggested_manifest["artifacts"]["gpio_led_table"], "GPIO_R0A_20250730.xlsx")
            self.assertEqual(len(result.suggested_manifest["artifacts"]["schematic_pdfs"]), 2)
            self.assertEqual(result.suggested_manifest["artifacts"]["public_ref_dts"], "968575REF1.dts")
            self.assertEqual(len(result.dts_files), 1)

            manifest_path = bootstrap_manifest(folder)
            self.assertTrue(manifest_path.exists())

            manifest = load_manifest(folder)
            self.assertEqual(manifest.project, "BGW720")
            self.assertEqual(manifest.family, "bcm68575")
            self.assertEqual(manifest.artifacts["public_ref_dts"], "968575REF1.dts")


if __name__ == "__main__":
    unittest.main()
